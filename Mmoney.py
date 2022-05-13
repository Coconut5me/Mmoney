import tkinter
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter.ttk import *
import time
import tkinter.ttk as exTk

import openpyxl
import pprint

import datetime
realDay = datetime.datetime.now().day
realMonth = datetime.datetime.now().month
realYear = datetime.datetime.now().year



wb = openpyxl.load_workbook('./Report.xlsx')
sheet = wb['Sheet1']

class Phieu_nhap:
    "Input money!"
    def __init__(self, type, day, month, year, source, src_desc, note, money):
        self.type = type
        self.day = day
        self.month = month
        self.year = year
        self.source = source
        self.src_desc = src_desc
        self.note = note
        self.money = money
    def NhapFile(self):
        row = sheet.max_row
        sheet['A'+str(row+1)] = self.type
        sheet['B'+str(row+1)] = self.day
        sheet['C'+str(row+1)] = self.month
        sheet['D'+str(row+1)] = self.year
        sheet['E'+str(row+1)] = self.source
        sheet['F'+str(row+1)] = self.src_desc
        sheet['G'+str(row+1)] = self.note
        sheet['H'+str(row+1)] = self.money
        wb.save('Report.xlsx')
    def XuatFile(self):
        if (self.type == "Thu"):
            tv.insert(parent = '', index = 'end', values=(str(self.day), str(self.type), str(self.src_desc), str(self.note), '+' + str(self.money)), tags = 'Thu')
        else:
            tv.insert(parent = '', index = 'end', values=(str(self.day), str(self.type), str(self.src_desc), str(self.note), '-' + str(self.money)), tags = 'Chi')




def changeOption():
    if(opion_label['text'] == 'Thu Nhập'):
        opion_label['text'] = 'Chi Tiêu'
        dm['values'] = ('Cần thiết', 'Đầu tư - tiết kiệm', 'Giải trí')
        dm.current(0)
    else:
        opion_label['text'] = 'Thu Nhập'
        dm['values'] = ('Lương', 'Phụ cấp', 'Khác')
        dm.current(0)

def nhap():
    if opion_label['text'] == 'Thu Nhập':
        x = Phieu_nhap('Thu', int(Day.get()), int(Month.get()), int(Year.get()), dm.current()+1, dm.get(), gc.get(), int(st.get()))
    else:
        x = Phieu_nhap('Chi', int(Day.get()), int(Month.get()), int(Year.get()), dm.current()+1, dm.get(), gc.get(), int(st.get()))
    x.NhapFile()
    clear()
    messagebox.showinfo(title='Chúc mừng!', message='Bạn đã nhập thành công!')

def clear():
    gc.delete(0, END)
    st.delete(0, END)

def show():
    removeAll()
    a = []
    i = 2
    row = sheet.max_row
    while (i <= row):
        if ((sheet['C'+str(i)].value == int(getMonth.get())) and (sheet['D'+str(i)].value == int(getYear.get()))):
            a.append(i)
        i = i + 1
    for j in range(len(a)):
        x = Phieu_nhap(sheet['A'+str(a[j])].value, sheet['B'+str(a[j])].value, sheet['C'+str(a[j])].value, sheet['D'+str(a[j])].value, sheet['E'+str(a[j])].value, sheet['F'+str(a[j])].value, sheet['G'+str(a[j])].value, sheet['H'+str(a[j])].value)
        x.XuatFile()

def removeAll():
    for record in tv.get_children():
        tv.delete(record)

def TongThu():
    x = Month_result.get()
    kq = 0
    row = sheet.max_row
    i = 2
    while (i <= row):
        if ((sheet['C'+str(i)].value == int(x)) and (sheet['A'+str(i)].value == "Thu") and (sheet['D'+str(i)].value == realYear)):
            kq = kq + sheet['H'+str(i)].value
        i = i + 1
    return kq

def TongChi():
    x = Month_result.get()
    kq = 0
    row = sheet.max_row
    i = 2
    while (i <= row):
        if ((sheet['C'+str(i)].value == int(x)) and (sheet['A'+str(i)].value == "Chi") and (sheet['D'+str(i)].value == realYear)):
            kq = kq + sheet['H'+str(i)].value
        i = i + 1
    return kq

def show_tk():
    lbl_thu = Label(tab3, text='Bạn đã thu: ' + str(TongThu()) + ' VNĐ', font='Times 15')
    lbl_thu.place(x=60, y=150)
    lbl_chi = Label(tab3, text='Bạn đã chi: ' + str(TongChi()) + ' VNĐ', font='Times 15')
    lbl_chi.place(x=60, y=200)
    XuLyChi()

    du = TongThu() - TongChi()
    if (du < 0):
        lbl_du = Label(tab3, text='Tháng này bạn đã nợ: ' + str(-du) + ' VNĐ', background='green', foreground='white',font='Times 15 bold')
        lbl_du.place(x=60, y=250)

    else:
        lbl_du = Label(tab3, text='Tháng này bạn còn dư: ' + str(du) + ' VNĐ', background='green', foreground='white',font='Times 15 bold')
        lbl_du.place(x=60, y=250)
        if (du > 0):
            lbl_xuly = Label(tab3, text='Bạn có muốn đầu tư tiết kiệm?', font='Times 15 bold')
            lbl_xuly.place(x=60, y=300)
            yes_button = Button(tab3, text='Yes',command=lambda: [yes_button.grid_forget(), no_button.grid_forget(), lbl_xuly.destroy(),lbl_du.destroy(), yes()], width=10)
            no_button = Button(tab3, text='No',command=lambda: [yes_button.grid_forget(), no_button.grid_forget(), lbl_xuly.destroy(),lbl_du.destroy(), no()], width=10)
            yes_button.place(x=380, y=300)
            no_button.place(x=480, y=300)

def yes():
    x = Phieu_nhap('Chi', 32, int(Month_result.get()), realYear, 2, 'Đầu tư - tiết kiệm', "Đầu tư từ tiền dư", TongThu() - TongChi())
    x.NhapFile()
    show_tk()

def no():
    y = Phieu_nhap('Thu', 0, int(Month_result.get()) + 1, realYear, 2, 'Khác', "Tiền dư từ tháng trước", TongThu() - TongChi())
    y.NhapFile()
    x = Phieu_nhap('Chi', 32, int(Month_result.get()), realYear, 2, 'Đầu tư - tiết kiệm', "Đầu tư vào tháng sau", TongThu() - TongChi())
    x.NhapFile()
    show_tk()

def XuLyChi():
    x = Month_result.get()
    chi1 = 0;
    chi2 = 0;
    chi3 = 0;
    tc = TongChi()
    row = sheet.max_row
    if (tc > 0):
        i = 2
        while (i <= row):
            if ((sheet['C'+str(i)].value == int(x)) and (sheet['A'+str(i)].value == "Chi") and (sheet['D'+str(i)].value == realYear)):
                if (sheet['E' + str(i)].value == 1):
                    chi1 = chi1 + sheet['H'+str(i)].value 
                elif (sheet['E' + str(i)].value == 2):
                    chi2 = chi2 + sheet['H'+str(i)].value
                else:
                    chi3 = chi3 + sheet['H'+str(i)].value

            i = i + 1

        if ((chi1/TongChi()) >= 0.47) and ((chi1/tc) <= 0.53) and ((chi2/tc) >= 0.17) and ((chi2/tc) <= 0.23) and ((chi3/tc) >= 0.27) and ((chi3/tc) <= 0.33):
            messagebox.showinfo("Amazing!", "Bạn có kế hoạch chi tiêu tuyệt vời!")
        else:
            if ((chi1 / tc) > 0.53) and ((chi3 / tc) > 0.33) :
                 messagebox.showwarning("Bạn nên điều chỉnh lại kế hoạch chi tiêu bản thân và bớt ăn chơi đi!")
            elif ((chi1 / tc) > 0.53) :
                messagebox.showwarning("Cảnh báo!", "Bạn cần có kế hoạch chi tiêu tốt hơn!")
            elif ((chi3 / tc) > 0.33):
                messagebox.showwarning("Cảnh báo!", "Bạn càn bớt ăn xài lại!")
            else:
                messagebox.showerror("Oh no oh no...", "Bạn cần có kế hoạch chi tiêu tối ưu hơn!")

        ct= Label(tab3, text = " cần thiết",width =8,background='red', foreground="white", font = 'Times 15 bold').place(x=40,y=90)
        dttk= Label(tab3, text = "   đt - tk  ",width =8,background='#ffd700', foreground="white",font = 'Times 15 bold').place(x=140,y=90)
        gt= Label(tab3, text = "  giải trí ",width =8,background='#1e90ff', foreground="white",font = 'Times 15 bold').place(x=240,y=90)

        Label(tab3, text = "BIỂU ĐỒ CHI TIÊU",font = 'Times 12 bold').place(x=380,y=60)
        canvas = Canvas(tab3, width = 500, height = 500)
        canvas.place(x=370,y=100)

        canvas.create_arc((2,2,150,150), fill = "red", outline = "red", start = 0, extent = chi1*360/tc)
        canvas.create_arc((2,2,150,150), fill = "#ffd700", outline = "#ffd700", start = chi1*360/tc, extent = chi2*360/tc)
        canvas.create_arc((2,2,150,150), fill = "#1e90ff", outline = "#1e90ff", start = chi1*360/tc + chi2*360/tc, extent = chi3*360/tc)

win = Tk()
win.title("Mmoney - CHƯƠNG TRÌNH QUẢN LÝ THU CHI")

scrH = win.winfo_screenheight()
scrW = win.winfo_screenwidth()
win.geometry('600x400+%d+%d' %(scrW/2-300,scrH/2-200)) #%d+%d' %(scrW/2-300,scrH/2-200): khi chương trình xuất hiện sẽ nằm chính giữa scr
win.resizable(width = False, height = False) #không cho resize lại nếu để = True thì ng dùng có thể resize lại

tab_control=ttk.Notebook(win)
tab1=ttk.Frame(tab_control)
tab2=ttk.Frame(tab_control)
tab3=ttk.Frame(tab_control)


tab_control.add(tab1,text = ' '*5 + 'NHẬP CHI TIÊU' + ' '*5)
tab_control.add(tab2,text = ' '*5 + 'CHI TIẾT' + ' '*5)
tab_control.add(tab3,text = ' '*5 + 'SỐ DƯ VÀ LỜI KHUYÊN' + ' '*5)
tab_control.pack(expand = 1, fill = 'both')


#tab1
opion_label = Label(tab1, text = 'Thu Nhập',background = '#7fffd4', font = 'Times 15 bold')
opion_label.place(x=200,y=20)
option_button=Button(tab1,text='Thay đổi', command = changeOption,)
option_button.place(x=320,y=20)

lbl_d = Label(tab1, text = 'Ngày ', font = 'Times 15')
lbl_d.place(x=45,y=70)
Day = Combobox(tab1, width = 3, font = 'Times 15',state = 'readonly')
Day['values'] = (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31)
Day.current(realDay-1)
Day.place(x=130,y=70)


lbl_m = Label(tab1, text = 'Tháng ', font = 'Times 15')
lbl_m.place(x=220,y=70)
Month = Combobox(tab1, width = 3, font = 'Times 15',state = 'readonly')
Month['values'] = (1,2,3,4,5,6,7,8,9,10,11,12)
Month.current(realMonth-1)
Month.place(x=305,y=70)

lbl_m = Label(tab1, text = 'Năm ', font = 'Times 15')
lbl_m.place(x=395,y=70)
Year = Combobox(tab1, width = 5, font = 'Times 15',state = 'readonly')
Year['values'] = (2021,2022)
Year.current(1)
Year.place(x=480,y=70)

lbl_dm = Label(tab1, text = 'Danh mục ', font = 'Times 15')
lbl_dm.place(x=45,y=130)
dm = Combobox(tab1, width = 20, font = 'Times 15', state = 'readonly')
dm['values'] = ('Lương', 'Phụ cấp', 'Khác')
dm.current(0)
dm.place(x=150,y=130)

lbl_gc = Label(tab1, text = 'Ghi chú ', font = 'Times 15')
lbl_gc.place(x=45,y=190)
gc = Entry(tab1, width = 40, font = 'Times 15')
gc.place(x=150,y=190)

lbl_st = Label(tab1, text = 'Số tiền ', font = 'Times 15')
lbl_st.place(x=45,y=250)
st = Entry(tab1, width = 40, font = 'Times 15')
st.place(x=150,y=250)

nhap_button = Button(tab1, text = 'NHẬP', command = nhap, width = 20)
nhap_button.place(x=260,y=320)

#tab2
khoangcach = Label(tab2, text = 'Bạn muốn xem thu chi tháng nào?', font = 'Times 15 italic')
khoangcach.place(x=20,y=40)
getMonth = Combobox(tab2, width = 3, font = 'Times 15', state = 'readonly')
getMonth['values'] = (1,2,3,4,5,6,7,8,9,10,11,12)
getMonth.current(realMonth-1)
getMonth.place(x=320,y=40)
getYear = Combobox(tab2, width = 5, font = 'Times 15', state = 'readonly')
getYear['values'] = (2021, 2022)
getYear.current(1)
getYear.place(x=380,y=40)
show_button = Button(tab2, text = 'Xem', command = show, width = 10)
show_button.place(x=500,y=40)

#tạo bảng
tv = ttk.Treeview(tab2)
tv['columns']=('Ngày', 'Loại', 'Danh mục', 'Ghi chú', 'Biến động')
tv.column('#0', width=0, stretch=NO)
tv.column('Ngày', anchor=CENTER, width=60)
tv.column('Loại', anchor=CENTER, width=60)
tv.column('Danh mục', anchor=CENTER, width=100)
tv.column('Ghi chú', anchor=CENTER, width=180)
tv.column('Biến động', anchor=CENTER, width=100)

tv.heading('#0', text='', anchor=CENTER)
tv.heading('Ngày', text='Ngày', anchor=CENTER)
tv.heading('Loại', text='Loại', anchor=CENTER)
tv.heading('Danh mục', text='Danh mục', anchor=CENTER)
tv.heading('Ghi chú', text='Ghi chú', anchor=CENTER)
tv.heading('Biến động', text='Biến động', anchor=CENTER)
tv.tag_configure('Thu', background = '#00ff00')
tv.tag_configure('Chi',foreground="white", background = 'red')
tv.place(x=50,y=100)

#tab3
lbl_th = Label(tab3, text = 'Tháng ', background = '#7fffd4',font = 'Times 15 bold')
lbl_th.place(x=60,y=20)
Month_result = Combobox(tab3, width = 3, font = 'Times 15',state = 'readonly')
Month_result['values'] = (1,2,3,4,5,6,7,8,9,10,11,12)
Month_result.current(realMonth-1)
Month_result.place(x=200,y=20)
tk_button = Button(tab3, text = 'Xem', command = show_tk, width = 10)
tk_button.place(x=350,y=20)

win.mainloop()